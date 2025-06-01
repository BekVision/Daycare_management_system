# reports/utils.py
import io
import csv
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import List, Dict, Any, Optional
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.conf import settings


class ReportExporter:
    """Hisobotlarni turli formatlarda eksport qilish"""

    @staticmethod
    def export_to_csv(data: List[Dict], filename: str, headers: List[str] = None) -> HttpResponse:
        """CSV formatida eksport"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'

        if not data:
            return response

        writer = csv.writer(response)

        # Headers
        if headers:
            writer.writerow(headers)
        elif data:
            writer.writerow(data[0].keys())

        # Data
        for row in data:
            writer.writerow(row.values())

        return response

    @staticmethod
    def export_to_excel(data: List[Dict], filename: str, headers: List[str] = None) -> HttpResponse:
        """Excel formatida eksport (openpyxl kerak)"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl kutubxonasi o'rnatilmagan")

        # Workbook yaratish
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hisobot"

        if not data:
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            wb.save(response)
            return response

        # Headers
        header_row = headers if headers else list(data[0].keys())
        ws.append(header_row)

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data
        for row in data:
            ws.append(list(row.values()))

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        wb.save(response)
        return response

    @staticmethod
    def export_to_pdf(context: Dict, template_name: str, filename: str) -> HttpResponse:
        """PDF formatida eksport (reportlab yoki weasyprint kerak)"""
        try:
            from weasyprint import HTML, CSS
            from django.template.loader import render_to_string
        except ImportError:
            raise ImportError("weasyprint kutubxonasi o'rnatilmagan")

        # HTML yaratish
        html_string = render_to_string(template_name, context)

        # CSS
        css_string = '''
        @page {
            size: A4;
            margin: 2cm;
        }
        body {
            font-family: Arial, sans-serif;
            font-size: 12px;
            line-height: 1.4;
        }
        .header {
            text-align: center;
            margin-bottom: 20px;
            border-bottom: 2px solid #366092;
            padding-bottom: 10px;
        }
        .table {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 20px;
        }
        .table th, .table td {
            border: 1px solid #ddd;
            padding: 8px;
            text-align: left;
        }
        .table th {
            background-color: #366092;
            color: white;
            font-weight: bold;
        }
        .text-center { text-align: center; }
        .text-right { text-align: right; }
        .footer {
            position: fixed;
            bottom: 0;
            width: 100%;
            text-align: center;
            font-size: 10px;
            border-top: 1px solid #ddd;
            padding-top: 5px;
        }
        '''

        # PDF yaratish
        html = HTML(string=html_string, base_url=settings.BASE_DIR)
        css = CSS(string=css_string)
        pdf = html.write_pdf(stylesheets=[css])

        # Response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        response.write(pdf)
        return response


class ReportCalculator:
    """Hisobot hisoblash yordamchi funktsiyalari"""

    @staticmethod
    def calculate_efficiency(served: int, planned: int) -> Optional[float]:
        """Samaradorlik foizini hisoblash"""
        if planned <= 0:
            return None
        return (served / planned) * 100

    @staticmethod
    def calculate_waste_percentage(waste: float, total_used: float) -> Optional[float]:
        """Chiqindi foizini hisoblash"""
        if total_used <= 0:
            return None
        return (waste / total_used) * 100

    @staticmethod
    def calculate_cost_per_portion(total_cost: Decimal, portions: int) -> Optional[Decimal]:
        """Porsiya uchun xarajatni hisoblash"""
        if portions <= 0:
            return None
        return total_cost / portions

    @staticmethod
    def calculate_growth_rate(current: float, previous: float) -> Optional[float]:
        """O'sish sur'atini hisoblash"""
        if previous <= 0:
            return None
        return ((current - previous) / previous) * 100

    @staticmethod
    def get_trend_data(queryset, date_field: str, value_field: str, days: int = 30) -> Dict[str, List]:
        """Trend ma'lumotlarini olish"""
        end_date = date.today()
        start_date = end_date - timedelta(days=days)

        data = list(queryset.filter(
            **{f'{date_field}__range': [start_date, end_date]}
        ).order_by(date_field).values(date_field, value_field))

        dates = []
        values = []

        for item in data:
            dates.append(item[date_field].strftime('%d.%m'))
            values.append(float(item[value_field]) if item[value_field] else 0)

        return {
            'dates': dates,
            'values': values
        }


class ReportFormatter:
    """Hisobot formatlovchi"""

    @staticmethod
    def format_currency(amount: Decimal, currency: str = 'so\'m') -> str:
        """Pul summasini formatlash"""
        if amount is None:
            return '-'
        return f"{amount:,.0f} {currency}"

    @staticmethod
    def format_percentage(value: float) -> str:
        """Foizni formatlash"""
        if value is None:
            return '-'
        return f"{value:.1f}%"

    @staticmethod
    def format_weight(weight: float, unit: str = 'kg') -> str:
        """Og'irlikni formatlash"""
        if weight is None:
            return '-'
        return f"{weight:.1f} {unit}"

    @staticmethod
    def format_date_range(start_date: date, end_date: date) -> str:
        """Sana oralig'ini formatlash"""
        if start_date == end_date:
            return start_date.strftime('%d.%m.%Y')
        return f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"

    @staticmethod
    def get_performance_badge(percentage: float) -> Dict[str, str]:
        """Ishlash ko'rsatkichi uchun badge"""
        if percentage >= 90:
            return {'class': 'bg-success', 'text': 'A\'lo'}
        elif percentage >= 80:
            return {'class': 'bg-primary', 'text': 'Yaxshi'}
        elif percentage >= 70:
            return {'class': 'bg-warning', 'text': 'O\'rta'}
        else:
            return {'class': 'bg-danger', 'text': 'Yomon'}

    @staticmethod
    def get_waste_badge(percentage: float) -> Dict[str, str]:
        """Chiqindi ko'rsatkichi uchun badge"""
        if percentage < 3:
            return {'class': 'bg-success', 'text': 'A\'lo'}
        elif percentage < 7:
            return {'class': 'bg-primary', 'text': 'Yaxshi'}
        elif percentage < 15:
            return {'class': 'bg-warning', 'text': 'O\'rta'}
        else:
            return {'class': 'bg-danger', 'text': 'Yomon'}


class ReportValidator:
    """Hisobot ma'lumotlarini tekshirish"""

    @staticmethod
    def validate_date_range(start_date: date, end_date: date) -> List[str]:
        """Sana oralig'ini tekshirish"""
        errors = []

        if start_date > end_date:
            errors.append("Boshlanish sanasi tugash sanasidan katta bo'lmasligi kerak")

        if end_date > date.today():
            errors.append("Tugash sanasi bugungi kundan katta bo'lmasligi kerak")

        # Maksimal oraliq (masalan, 2 yil)
        max_days = 730
        if (end_date - start_date).days > max_days:
            errors.append(f"Maksimal oraliq {max_days} kun")

        return errors

    @staticmethod
    def validate_report_data(data: Dict) -> List[str]:
        """Hisobot ma'lumotlarini tekshirish"""
        errors = []

        # Majburiy maydonlar
        required_fields = ['report_date', 'total_portions_served']
        for field in required_fields:
            if field not in data or data[field] is None:
                errors.append(f"{field} maydoni majburiy")

        # Musbat qiymatlar
        positive_fields = ['total_portions_served', 'total_cost']
        for field in positive_fields:
            if field in data and data[field] is not None and data[field] < 0:
                errors.append(f"{field} musbat bo'lishi kerak")

        # Foiz qiymatlari (0-100)
        percentage_fields = ['efficiency_percentage', 'waste_percentage']
        for field in percentage_fields:
            if field in data and data[field] is not None:
                if data[field] < 0 or data[field] > 100:
                    errors.append(f"{field} 0-100 orasida bo'lishi kerak")

        return errors